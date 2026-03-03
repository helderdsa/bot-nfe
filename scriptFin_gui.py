import os
import sys
import queue
import threading
import time
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import customtkinter as ctk
from tkinter import messagebox, filedialog
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# ============================================================
# Lógica do bot (adaptada para rodar em thread separada)
# ============================================================

def run_bot(start_date: str, end_date: str, cancel_event: threading.Event, log: callable, resume_path: str = None, categoria_filtro: str = None):
    """
    Executa o bot de emissão de NF-e.

    Parâmetros
    ----------
    start_date       : str  – data inicial no formato YYYY-MM-DD
    end_date         : str  – data final   no formato YYYY-MM-DD
    cancel_event     : threading.Event – setado externamente para cancelar
    log              : callable(str) – função para enviar mensagens ao log da GUI
    resume_path      : str  – caminho para o XLSX de relatório anterior (opcional)
    categoria_filtro : str  – categoria para filtrar (None = todas)
    """

    options = Options()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument("--start-maximized")
    options.add_experimental_option(
        "prefs",
        {"profile.default_content_setting_values.automatic_downloads": 1}
    )
    options.add_argument("--ignore-certificate-errors")

    driver = None
    transactions = []
    ids_emitidas = []
    ids_erro = []
    ids_puladas = []

    # ===== RETOMADA POR RELATÓRIO ANTERIOR =====
    # Colunas do relatório: 0=ID, 9=Situação (SUCESSO | ERRO | IGNORADO)
    ids_ja_emitidas: set = set()   # SUCESSO  – pular
    ids_ignoradas: set = set()     # IGNORADO – pular
    ids_para_retentar: set = set() # ERRO     – retentar
    usando_relatorio = bool(resume_path)

    if resume_path:
        try:
            wb_resume = openpyxl.load_workbook(resume_path)
            ws_resume = wb_resume.active
            for row in ws_resume.iter_rows(min_row=2, values_only=True):
                tid      = row[0] if len(row) > 0 else None
                situacao = row[9] if len(row) > 9 else None
                if tid is None:
                    continue
                if situacao == "SUCESSO":
                    ids_ja_emitidas.add(tid)
                elif situacao == "ERRO":
                    ids_para_retentar.add(tid)
                elif situacao == "IGNORADO":
                    ids_ignoradas.add(tid)
            log("✓ Relatório anterior carregado:")
            log(f"  - ✅ SUCESSO  (serão puladas):  {len(ids_ja_emitidas)}")
            log(f"  - ❌ ERRO     (serão retentadas): {len(ids_para_retentar)}")
            log(f"  - ⏭ IGNORADO (serão puladas):  {len(ids_ignoradas)}")
        except Exception as exc:
            log(f"✗ Erro ao ler o relatório anterior: {exc}")
            return

    try:
        # ===== INICIALIZAÇÃO DO CHROME =====
        log("Iniciando o Chrome...")
        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 20)
        log("✓ Chrome iniciado com sucesso!")

        # ===== REQUISIÇÃO À API =====
        log("Fazendo requisição à API...")

        api_base_url = (
            f"https://app.advbox.com.br/api/v1/transactions"
            f"?date_payment_start={start_date}&date_payment_end={end_date}"
        )

        session = requests.Session()
        headers = {
            "Authorization": "Bearer 2gAucrPVxikEyTeHNj0QIhLQci2NE9u2hZTndQPV2D1E96J2RBfEaaVfG2Xh",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
        }

        LIMIT = 1000
        offset = 0
        total_count = None

        while True:
            if cancel_event.is_set():
                log("⚠ Cancelamento solicitado antes de concluir a busca na API.")
                break

            api_url = f"{api_base_url}&offset={offset}&limit={LIMIT}"
            response = session.get(api_url, headers=headers)
            log(f"Status Code: {response.status_code} (offset={offset})")

            if response.status_code != 200:
                log(f"✗ Erro na API. Status: {response.status_code}")
                log(f"✗ Resposta: {response.text}")
                break

            data_json = response.json()

            if total_count is None:
                total_count = data_json.get("totalCount", 0)
                log(f"✓ Total (totalCount): {total_count}")

            page_data = data_json.get("data", [])
            transactions.extend(page_data)
            log(f"✓ Página carregada: {len(page_data)} registros (acumulado: {len(transactions)})")

            if len(transactions) >= total_count or len(page_data) == 0:
                break

            offset += LIMIT

        log(f"✓ Total de transações encontradas: {len(transactions)}")

        # ===== FILTRO DE CATEGORIA =====
        if categoria_filtro and categoria_filtro != "TODAS":
            transactions = [t for t in transactions if t.get("category") == categoria_filtro]
            log(f"✓ Transações após filtro '{categoria_filtro}': {len(transactions)}")

        # ===== PROCESSAMENTO =====
        if len(transactions) == 0:
            log("⚠ Nenhuma transação encontrada para o período informado.")
            return

        driver.get("https://www.nfse.gov.br/EmissorNacional/Login?ReturnUrl=%2fEmissorNacional")
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "img-certificado")))
        time.sleep(0.5)
        driver.find_element(By.CLASS_NAME, "img-certificado").click()

        log("\n===== Processando transações =====")
        index = 0
        max_tentativas = 5

        while index < len(transactions):
            if cancel_event.is_set():
                log("⚠ Cancelamento solicitado. Salvando relatório...")
                break

            transaction = transactions[index]

            tid = transaction.get("id")

            # ── Filtro de retomada ──────────────────────────────────────
            if usando_relatorio:
                if tid in ids_ja_emitidas:
                    log(f"\n⏭ ID {tid} - já emitida com sucesso anteriormente. Pulando.")
                    index += 1
                    continue
                # ERRO, IGNORADO e IDs novas são processadas normalmente
            # ───────────────────────────────────────────────────────────

            if transaction.get("entry_type") == "expense":
                log(f"\n--- Transação {index + 1}/{len(transactions)} ---")
                log(f"⏭ Pulando ID {tid} - despesa não gera NF.")
                ids_puladas.append(tid)
                index += 1
                continue

            if transaction.get("category") == "REEMBOLSO DE CUSTO POR CLIENTES":
                log(f"\n--- Transação {index + 1}/{len(transactions)} ---")
                log(f"⏭ Pulando ID {tid} - 'REEMBOLSO DE CUSTO POR CLIENTES' não gera NF.")
                ids_puladas.append(tid)
                index += 1
                continue

            tentativas = 0
            sucesso = False

            while tentativas < max_tentativas and not sucesso:
                if cancel_event.is_set():
                    log("⚠ Cancelamento solicitado durante tentativa. Parando.")
                    sucesso = True  # sai do loop de tentativas
                    continue

                try:
                    if tentativas > 0:
                        log(f"\n⚠ Tentativa {tentativas + 1}/{max_tentativas}")

                    log(f"\n--- Transação {index + 1}/{len(transactions)} ---")
                    log(f"ID: {transaction.get('id')}")
                    log(f"Tipo: {transaction.get('entry_type')}")
                    log(f"Data Pagamento: {transaction.get('date_payment')}")
                    log(f"Valor: {transaction.get('amount')}")
                    log(f"Nome: {transaction.get('name')}")
                    log(f"Categoria: {transaction.get('category')}")

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".btnAcesso")))
                    time.sleep(0.5)
                    driver.find_element(By.CSS_SELECTOR, ".btnAcesso").click()

                    wait.until(EC.presence_of_element_located((By.ID, "DataCompetencia")))
                    time.sleep(0.5)
                    driver.find_element(By.ID, "DataCompetencia").send_keys(transaction.get("date_payment", ""))

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "body")))
                    time.sleep(0.5)
                    driver.find_element(By.CSS_SELECTOR, "body").click()

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".radiobutton")))
                    time.sleep(0.5)
                    radio_options = driver.find_elements(By.CSS_SELECTOR, ".radiobutton")
                    radio_options[4].find_element(By.CSS_SELECTOR, "label").click()

                    wait.until(EC.presence_of_element_located((By.ID, "Tomador_Inscricao")))
                    time.sleep(0.5)
                    driver.find_element(By.ID, "Tomador_Inscricao").click()
                    time.sleep(0.5)
                    driver.find_element(By.ID, "Tomador_Inscricao").send_keys(
                        str(transaction.get("identification", ""))
                    )

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "body")))
                    time.sleep(0.5)
                    driver.find_element(By.CSS_SELECTOR, "body").click()

                    wait.until(EC.presence_of_element_located((By.ID, "btnAvancar")))
                    time.sleep(0.5)
                    driver.find_element(By.ID, "btnAvancar").click()

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".select2-selection")))
                    time.sleep(0.5)
                    driver.find_element(By.CSS_SELECTOR, ".select2-selection").click()

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input.select2-search__field")))
                    time.sleep(0.5)
                    driver.find_element(By.CSS_SELECTOR, "input.select2-search__field").send_keys("Acari")

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".select2-results__option--highlighted")))
                    time.sleep(0.5)
                    driver.find_element(By.CSS_SELECTOR, ".select2-results__option--highlighted").click()

                    wait.until(EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "span[aria-labelledby=select2-ServicoPrestado_CodigoTributacaoNacional-container]")
                    ))
                    time.sleep(0.5)
                    driver.find_element(
                        By.CSS_SELECTOR,
                        "span[aria-labelledby=select2-ServicoPrestado_CodigoTributacaoNacional-container]"
                    ).click()

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input.select2-search__field")))
                    time.sleep(0.5)
                    driver.find_element(By.CSS_SELECTOR, "input.select2-search__field").send_keys("171401")

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".select2-results__option--highlighted")))
                    time.sleep(0.5)
                    driver.find_element(By.CSS_SELECTOR, ".select2-results__option--highlighted").click()

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.radiobutton > label")))
                    time.sleep(0.5)
                    driver.find_element(By.CSS_SELECTOR, "div.radiobutton > label").click()

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#ServicoPrestado_Descricao")))
                    time.sleep(0.5)
                    driver.find_element(By.CSS_SELECTOR, "#ServicoPrestado_Descricao").click()
                    time.sleep(0.5)

                    category = transaction.get("category", "")
                    if category == "IMPLANTAÇÕES":
                        descricao = "Pagamento referente a implantação de letras, parcela " + str(transaction.get("description", ""))
                    elif category in ("PRECATÓRIOS", "ALVARÁS", "HONORÁRIOS DE SUCUMBÊNCIA"):
                        descricao = "Pagamento referente aos honorários contratuais, nº do processo: " + str(transaction.get("process_number", ""))
                    else:
                        descricao = "Pagamento referente a " + str(category)

                    driver.find_element(By.CSS_SELECTOR, "#ServicoPrestado_Descricao").send_keys(descricao)

                    wait.until(EC.presence_of_element_located((By.CLASS_NAME, "btn-primary")))
                    time.sleep(0.5)
                    driver.find_element(By.CLASS_NAME, "btn-primary").click()

                    wait.until(EC.presence_of_element_located((By.ID, "Valores_ValorServico")))
                    time.sleep(0.5)
                    valor = transaction.get("amount", "0")
                    valor_formatado = f"{float(valor):.2f}".replace(".", ",")
                    driver.find_element(By.ID, "Valores_ValorServico").send_keys(valor_formatado)

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".radiobutton")))
                    time.sleep(0.5)
                    radio_options = driver.find_elements(By.CSS_SELECTOR, ".radiobutton")
                    radio_options[0].find_element(By.CSS_SELECTOR, "label").click()
                    radio_options[2].find_element(By.CSS_SELECTOR, "label").click()
                    radio_options[6].find_element(By.CSS_SELECTOR, "label").click()

                    wait.until(EC.presence_of_element_located(
                        (By.ID, "TributacaoFederal_PISCofins_SituacaoTributaria_chosen")
                    ))
                    driver.find_elements(By.ID, "TributacaoFederal_PISCofins_SituacaoTributaria_chosen")[0].click()
                    driver.find_elements(By.ID, "TributacaoFederal_PISCofins_SituacaoTributaria_chosen")[0].find_element(
                        By.CSS_SELECTOR, "div > div > input"
                    ).send_keys("00")

                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div > div > ul > li")))
                    driver.find_elements(By.ID, "TributacaoFederal_PISCofins_SituacaoTributaria_chosen")[0].find_element(
                        By.CSS_SELECTOR, "div > div > ul > li"
                    ).click()

                    wait.until(EC.presence_of_element_located(
                        (By.ID, "TributacaoFederal_PISCofins_TipoRetencao_chosen")
                    ))
                    driver.find_elements(By.ID, "TributacaoFederal_PISCofins_TipoRetencao_chosen")[0].click()

                    wait.until(EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "div > ul > li[data-option-array-index='1']")
                    ))
                    driver.find_elements(By.ID, "TributacaoFederal_PISCofins_TipoRetencao_chosen")[0].find_element(
                        By.CSS_SELECTOR, "div > ul > li[data-option-array-index='1']"
                    ).click()
                    time.sleep(0.5)

                    wait.until(EC.presence_of_element_located((By.CLASS_NAME, "btn-primary")))
                    time.sleep(0.5)
                    driver.find_element(By.CLASS_NAME, "btn-primary").click()

                    wait.until(EC.presence_of_element_located((By.ID, "btnProsseguir")))
                    time.sleep(0.5)
                    driver.find_element(By.ID, "btnProsseguir").click()

                    wait.until(EC.presence_of_element_located((By.ID, "btnDownloadDANFSE")))
                    time.sleep(0.5)
                    driver.get("https://www.nfse.gov.br/EmissorNacional/Dashboard")

                    sucesso = True
                    ids_emitidas.append(transaction.get("id"))
                    log(f"✓ Transação {index + 1} processada com sucesso!")

                except Exception as e:
                    tentativas += 1
                    log(f"\n✗ Erro na transação {index + 1}: {e}")
                    log(f"✗ Tipo: {type(e).__name__}")
                    try:
                        driver.get("https://www.nfse.gov.br/EmissorNacional/Dashboard")
                        time.sleep(10)
                    except Exception:
                        pass

                    if tentativas >= max_tentativas:
                        log(f"✗ Máximo de tentativas atingido. Pulando transação {index + 1}.")
                        ids_erro.append(transaction.get("id"))
                        sucesso = True

            index += 1

    except requests.exceptions.RequestException as e:
        log(f"✗ Erro na requisição à API: {e}")
    except Exception as e:
        log(f"✗ Erro inesperado: {e}")
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

        # Gera relatório independente de como o bot terminou
        ids_processados = set(ids_emitidas) | set(ids_erro) | set(ids_puladas)
        ids_faltaram = [t.get("id") for t in transactions if t.get("id") not in ids_processados]

        set_emitidas = set(ids_emitidas)
        set_erro     = set(ids_erro)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório NF"

        # ── Cabeçalho ──────────────────────────────────────────────
        headers = [
            "ID", "Tipo", "Data Pagamento", "Valor", "Nome",
            "Categoria", "Identificação", "Descrição", "Nº Processo", "Situação",
        ]
        ws.append(headers)
        hdr_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Fills de situação ──────────────────────────────────────
        fill_sucesso  = PatternFill(start_color="1E6B3C", end_color="1E6B3C", fill_type="solid")  # verde
        fill_erro     = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # vermelho
        fill_ignorado = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # azul
        font_status   = Font(bold=True, color="FFFFFF")

        # ── Linhas de dados ────────────────────────────────────────
        for t in transactions:
            tid = t.get("id")
            if tid in set_emitidas:
                situacao, fill = "SUCESSO", fill_sucesso
            elif tid in set_erro:
                situacao, fill = "ERRO", fill_erro
            else:
                situacao, fill = "IGNORADO", fill_ignorado

            ws.append([
                tid,
                t.get("entry_type", ""),
                t.get("date_payment", ""),
                t.get("amount", ""),
                t.get("name", ""),
                t.get("category", ""),
                t.get("identification", ""),
                t.get("description", ""),
                t.get("process_number", ""),
                situacao,
            ])
            situacao_cell = ws.cell(row=ws.max_row, column=len(headers))
            situacao_cell.fill = fill
            situacao_cell.font = font_status
            situacao_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Largura automática das colunas ─────────────────────────
        for col in ws.columns:
            max_w = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 55)

        # Quando empacotado pelo PyInstaller, usa a pasta do .exe
        if getattr(sys, 'frozen', False):
            script_dir = os.path.dirname(sys.executable)
        else:
            script_dir = os.path.dirname(os.path.abspath(__file__))
        filename = os.path.join(
            script_dir,
            f"relatorio_nf_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        wb.save(filename)

        log("\n===== RELATÓRIO FINAL =====")
        log(f"✓ Planilha salva: {filename}")
        log(f"  - ✅ SUCESSO  (verde):    {len(ids_emitidas)}")
        log(f"  - ❌ ERRO     (vermelho): {len(ids_erro)}")
        log(f"  - ⏭ IGNORADO (azul):     {len(transactions) - len(ids_emitidas) - len(ids_erro)}")


# ============================================================
# Interface Gráfica
# ============================================================

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.title("Bot Emissor de NF-e")
        self.geometry("700x600")
        self.resizable(False, False)

        self._cancel_event: threading.Event | None = None
        self._bot_thread: threading.Thread | None = None
        self._log_queue: queue.Queue = queue.Queue()
        self._resume_path: str = ""

        self._build_ui()
        self._poll_log_queue()

    # ----------------------------------------------------------
    # Construção da interface
    # ----------------------------------------------------------

    def _build_ui(self):
        # Título
        ctk.CTkLabel(
            self, text="Bot Emissor de NF-e",
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(pady=(20, 4))

        ctk.CTkLabel(
            self, text="Emissão automática via nfse.gov.br",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        ).pack(pady=(0, 16))

        # Frame de datas
        date_frame = ctk.CTkFrame(self)
        date_frame.pack(padx=30, fill="x")

        ctk.CTkLabel(date_frame, text="Data Inicial:").grid(
            row=0, column=0, padx=(16, 8), pady=12, sticky="w"
        )
        self.entry_start, self._var_start = self._make_date_entry(
            date_frame, placeholder_text="dd/mm/aaaa", width=160
        )
        self.entry_start.grid(row=0, column=1, padx=(0, 24), pady=12)

        ctk.CTkLabel(date_frame, text="Data Final:").grid(
            row=0, column=2, padx=(0, 8), pady=12, sticky="w"
        )
        self.entry_end, self._var_end = self._make_date_entry(
            date_frame, placeholder_text="dd/mm/aaaa", width=160
        )
        self.entry_end.grid(row=0, column=3, padx=(0, 16), pady=12)

        # Frame de filtro de categoria
        cat_frame = ctk.CTkFrame(self)
        cat_frame.pack(padx=30, pady=(12, 12), fill="x")

        ctk.CTkLabel(cat_frame, text="Categoria:").grid(
            row=0, column=0, padx=(16, 8), pady=12, sticky="w"
        )
        self._var_categoria = ctk.StringVar(value="TODAS")
        ctk.CTkOptionMenu(
            cat_frame,
            variable=self._var_categoria,
            values=["TODAS", "IMPLANTAÇÕES", "PRECATÓRIOS", "ALVARÁS"],
            width=220,
        ).grid(row=0, column=1, padx=(0, 16), pady=12, sticky="w")

        # Frame de seleção de relatório anterior
        file_frame = ctk.CTkFrame(self)
        file_frame.pack(padx=30, pady=(0, 4), fill="x")

        ctk.CTkLabel(file_frame, text="Relatório anterior:").grid(
            row=0, column=0, padx=(16, 8), pady=10, sticky="w"
        )
        self.entry_resume = ctk.CTkEntry(
            file_frame, width=360, placeholder_text="Opcional – selecione o .xlsx gerado na última execução"
        )
        self.entry_resume.grid(row=0, column=1, padx=(0, 8), pady=10)
        ctk.CTkButton(
            file_frame, text="Procurar", width=90, command=self._on_browse
        ).grid(row=0, column=2, padx=(0, 16), pady=10)
        ctk.CTkButton(
            file_frame, text="✕", width=32, fg_color="#555", hover_color="#333",
            command=self._on_clear_resume
        ).grid(row=0, column=3, padx=(0, 8), pady=10)

        # Botões
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(pady=12)

        self.btn_start = ctk.CTkButton(
            btn_frame, text="▶  Iniciar",
            width=140, command=self._on_start
        )
        self.btn_start.pack(side="left", padx=8)

        self.btn_cancel = ctk.CTkButton(
            btn_frame, text="✖  Cancelar",
            width=140, fg_color="#c0392b", hover_color="#922b21",
            state="disabled", command=self._on_cancel
        )
        self.btn_cancel.pack(side="left", padx=8)

        # Área de log
        ctk.CTkLabel(self, text="Log de execução:", anchor="w").pack(
            padx=30, fill="x"
        )
        self.log_box = ctk.CTkTextbox(self, height=280, font=ctk.CTkFont(family="Courier", size=11))
        self.log_box.pack(padx=30, pady=(4, 20), fill="both", expand=True)
        self.log_box.configure(state="disabled")

    # ----------------------------------------------------------
    # Cria campo de data com auto-formatação dd/mm/aaaa
    # ----------------------------------------------------------

    def _make_date_entry(self, parent, **kwargs):
        var = ctk.StringVar()
        entry = ctk.CTkEntry(parent, textvariable=var, **kwargs)
        trace_id = [None]

        def on_change(*_):
            current = var.get()
            digits = "".join(c for c in current if c.isdigit())[:8]

            if len(digits) <= 2:
                formatted = digits
            elif len(digits) <= 4:
                formatted = digits[:2] + "/" + digits[2:]
            else:
                formatted = digits[:2] + "/" + digits[2:4] + "/" + digits[4:]

            if formatted != current:
                # Remove trace para evitar recursão
                var.trace_remove("write", trace_id[0])
                var.set(formatted)
                trace_id[0] = var.trace_add("write", on_change)
                entry.after(0, lambda: entry.icursor("end"))

        trace_id[0] = var.trace_add("write", on_change)
        return entry, var

    # ----------------------------------------------------------
    # Converte dd/mm/aaaa → aaaa-mm-dd (formato da API)
    # ----------------------------------------------------------

    def _on_browse(self):
        path = filedialog.askopenfilename(
            title="Selecionar relatório anterior",
            filetypes=[("Planilha Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        if path:
            self._resume_path = path
            self.entry_resume.delete(0, "end")
            self.entry_resume.insert(0, path)

    def _on_clear_resume(self):
        self._resume_path = ""
        self.entry_resume.delete(0, "end")

    @staticmethod
    def _to_api_date(value: str) -> str:
        return datetime.strptime(value.strip(), "%d/%m/%Y").strftime("%Y-%m-%d")

    # ----------------------------------------------------------
    # Handlers de botão
    # ----------------------------------------------------------

    def _on_start(self):
        start_raw = self.entry_start.get().strip()
        end_raw = self.entry_end.get().strip()

        if not start_raw or not end_raw:
            messagebox.showwarning("Campos obrigatórios", "Preencha a data inicial e a data final.")
            return

        try:
            start = self._to_api_date(start_raw)
            end   = self._to_api_date(end_raw)
        except ValueError:
            messagebox.showerror("Formato inválido", "Use o formato dd/mm/aaaa nas datas.")
            return

        self._clear_log()
        self._cancel_event = threading.Event()
        self.btn_start.configure(state="disabled")
        self.btn_cancel.configure(state="normal")

        resume = self.entry_resume.get().strip() or None
        categoria = self._var_categoria.get()

        self._bot_thread = threading.Thread(
            target=self._run_bot_thread,
            args=(start, end, resume, categoria),
            daemon=True
        )
        self._bot_thread.start()

    def _on_cancel(self):
        if self._cancel_event:
            self._cancel_event.set()
        self.btn_cancel.configure(state="disabled", text="Cancelando…")

    # ----------------------------------------------------------
    # Thread do bot
    # ----------------------------------------------------------

    def _run_bot_thread(self, start: str, end: str, resume: str = None, categoria: str = None):
        try:
            run_bot(start, end, self._cancel_event, self._enqueue_log, resume_path=resume, categoria_filtro=categoria)
        except Exception as e:
            self._enqueue_log(f"✗ Erro fatal não tratado: {e}")
            self._enqueue_log(f"✗ Tipo: {type(e).__name__}")
        finally:
            self._enqueue_log("— Processo encerrado —")
            self.after(0, self._on_bot_finished)

    def _on_bot_finished(self):
        self.btn_start.configure(state="normal")
        self.btn_cancel.configure(state="disabled", text="✖  Cancelar")

    # ----------------------------------------------------------
    # Log assíncrono (thread-safe via queue)
    # ----------------------------------------------------------

    def _enqueue_log(self, message: str):
        self._log_queue.put(message)

    def _poll_log_queue(self):
        """Verifica a fila de log a cada 100 ms e exibe mensagens."""
        try:
            while True:
                msg = self._log_queue.get_nowait()
                self.log_box.configure(state="normal")
                self.log_box.insert("end", msg + "\n")
                self.log_box.see("end")
                self.log_box.configure(state="disabled")
        except queue.Empty:
            pass
        self.after(100, self._poll_log_queue)

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")


# ============================================================
# Ponto de entrada
# ============================================================

if __name__ == "__main__":
    app = App()
    app.mainloop()
